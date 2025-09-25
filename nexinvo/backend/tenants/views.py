from rest_framework import generics, viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db import models
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import pandas as pd
import openpyxl
from io import BytesIO
import json

from .models import Tenant, TenantMembership, Client, AuditLog
from .serializers import (
    TenantSerializer,
    TenantMembershipSerializer,
    ClientSerializer,
    AuditLogSerializer,
    TenantInviteSerializer
)


class TenantViewSet(viewsets.ModelViewSet):
    serializer_class = TenantSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # Users can only see tenants they are members of
        tenant_ids = user.tenant_memberships.values_list('tenant_id', flat=True)
        return Tenant.objects.filter(id__in=tenant_ids)

    def get_permissions(self):
        if self.action in ['create']:
            # Anyone can create a tenant
            return [permissions.IsAuthenticated()]
        elif self.action in ['update', 'partial_update', 'destroy']:
            # Only owners/admins can modify
            return [permissions.IsAuthenticated()]
        return super().get_permissions()

    def check_tenant_permission(self, tenant, required_roles=['owner', 'admin']):
        membership = TenantMembership.objects.filter(
            user=self.request.user,
            tenant=tenant,
            is_active=True,
            role__in=required_roles
        ).first()
        if not membership:
            return False
        return True

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if not self.check_tenant_permission(instance):
            return Response(
                {'error': 'You do not have permission to modify this tenant'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if not self.check_tenant_permission(instance, ['owner']):
            return Response(
                {'error': 'Only tenant owners can delete tenants'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def invite_member(self, request, pk=None):
        tenant = self.get_object()
        if not self.check_tenant_permission(tenant):
            return Response(
                {'error': 'You do not have permission to invite members'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = TenantInviteSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            membership = serializer.save()
            return Response(
                TenantMembershipSerializer(membership).data,
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        tenant = self.get_object()
        memberships = TenantMembership.objects.filter(tenant=tenant, is_active=True)
        serializer = TenantMembershipSerializer(memberships, many=True)
        return Response(serializer.data)


class ClientViewSet(viewsets.ModelViewSet):
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['client_type', 'state_code']
    search_fields = ['name', 'client_code', 'email', 'gstin']
    ordering_fields = ['name', 'client_code', 'created_at']

    def get_queryset(self):
        user = self.request.user
        # Users can only see clients from their tenant
        tenant_ids = user.tenant_memberships.filter(is_active=True).values_list('tenant_id', flat=True)
        return Client.objects.filter(tenant_id__in=tenant_ids)

    def check_client_permission(self, client=None, action='view'):
        user = self.request.user
        if client:
            tenant = client.tenant
        else:
            # For create actions, get tenant from user's membership
            membership = user.tenant_memberships.filter(is_active=True).first()
            if not membership:
                return False
            tenant = membership.tenant

        membership = TenantMembership.objects.filter(
            user=user,
            tenant=tenant,
            is_active=True
        ).first()

        if not membership:
            return False

        if action in ['create', 'update', 'delete']:
            return membership.role in ['owner', 'admin', 'finance_user', 'manager']
        return True  # Everyone can view

    def create(self, request, *args, **kwargs):
        if not self.check_client_permission(action='create'):
            return Response(
                {'error': 'You do not have permission to create clients'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        if not self.check_client_permission(instance, 'update'):
            return Response(
                {'error': 'You do not have permission to update this client'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if not self.check_client_permission(instance, 'delete'):
            return Response(
                {'error': 'You do not have permission to delete this client'},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['post'], url_path='upload-excel')
    def upload_excel(self, request):
        """Upload Excel file and create clients"""
        if not self.check_client_permission(action='create'):
            return Response(
                {'error': 'You do not have permission to create clients'},
                status=status.HTTP_403_FORBIDDEN
            )

        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']

        # Validate file type
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get user's tenant
            membership = request.user.tenant_memberships.filter(is_active=True).first()
            if not membership:
                return Response(
                    {'error': 'User is not a member of any tenant'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            tenant = membership.tenant

            # Read Excel file
            df = pd.read_excel(file, dtype=str)

            # Expected columns mapping
            column_mapping = {
                'Name': 'name',
                'Client Code': 'client_code',
                'Client Type': 'client_type',
                'GSTIN': 'gstin',
                'PAN': 'pan',
                'Email': 'email',
                'Phone': 'phone',
                'State Code': 'state_code',
                'Credit Terms (Days)': 'credit_terms_days',
                'Billing Address Line 1': 'billing_address_line1',
                'Billing Address Line 2': 'billing_address_line2',
                'Billing City': 'billing_city',
                'Billing State': 'billing_state',
                'Billing Pincode': 'billing_pincode',
                'Shipping Address Line 1': 'shipping_address_line1',
                'Shipping Address Line 2': 'shipping_address_line2',
                'Shipping City': 'shipping_city',
                'Shipping State': 'shipping_state',
                'Shipping Pincode': 'shipping_pincode'
            }

            # Validate required columns
            required_columns = ['Name', 'Client Code', 'Client Type']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process and create clients
            created_clients = []
            errors = []

            for index, row in df.iterrows():
                try:
                    # Skip empty rows
                    if pd.isna(row.get('Name')) or str(row.get('Name')).strip() == '':
                        continue

                    # Prepare client data
                    client_data = {
                        'tenant': tenant,
                        'name': str(row.get('Name', '')).strip(),
                        'client_code': str(row.get('Client Code', '')).strip(),
                        'client_type': str(row.get('Client Type', 'b2c')).lower(),
                        'gstin': str(row.get('GSTIN', '')).strip(),
                        'pan': str(row.get('PAN', '')).strip(),
                        'email': str(row.get('Email', '')).strip(),
                        'phone': str(row.get('Phone', '')).strip(),
                        'state_code': str(row.get('State Code', '')).strip(),
                        'pos_default': str(row.get('State Code', '')).strip(),
                        'credit_terms_days': 30,
                        'created_by': request.user,
                        'updated_by': request.user
                    }

                    # Handle credit terms
                    if pd.notna(row.get('Credit Terms (Days)')):
                        try:
                            client_data['credit_terms_days'] = int(float(row.get('Credit Terms (Days)')))
                        except (ValueError, TypeError):
                            client_data['credit_terms_days'] = 30

                    # Validate client type
                    if client_data['client_type'] not in ['b2b', 'b2c']:
                        client_data['client_type'] = 'b2c'

                    # Handle billing address
                    billing_address = {}
                    if pd.notna(row.get('Billing Address Line 1')):
                        billing_address['line1'] = str(row.get('Billing Address Line 1', '')).strip()
                    if pd.notna(row.get('Billing Address Line 2')):
                        billing_address['line2'] = str(row.get('Billing Address Line 2', '')).strip()
                    if pd.notna(row.get('Billing City')):
                        billing_address['city'] = str(row.get('Billing City', '')).strip()
                    if pd.notna(row.get('Billing State')):
                        billing_address['state'] = str(row.get('Billing State', '')).strip()
                    if pd.notna(row.get('Billing Pincode')):
                        billing_address['pincode'] = str(row.get('Billing Pincode', '')).strip()

                    client_data['billing_address'] = billing_address

                    # Handle shipping address
                    shipping_address = {}
                    if pd.notna(row.get('Shipping Address Line 1')):
                        shipping_address['line1'] = str(row.get('Shipping Address Line 1', '')).strip()
                    if pd.notna(row.get('Shipping Address Line 2')):
                        shipping_address['line2'] = str(row.get('Shipping Address Line 2', '')).strip()
                    if pd.notna(row.get('Shipping City')):
                        shipping_address['city'] = str(row.get('Shipping City', '')).strip()
                    if pd.notna(row.get('Shipping State')):
                        shipping_address['state'] = str(row.get('Shipping State', '')).strip()
                    if pd.notna(row.get('Shipping Pincode')):
                        shipping_address['pincode'] = str(row.get('Shipping Pincode', '')).strip()

                    client_data['shipping_address'] = shipping_address

                    # Check for duplicate client code within tenant
                    if Client.objects.filter(tenant=tenant, client_code=client_data['client_code']).exists():
                        errors.append(f"Row {index + 2}: Client with code '{client_data['client_code']}' already exists")
                        continue

                    # Create client
                    client = Client.objects.create(**client_data)
                    created_clients.append({
                        'id': str(client.id),
                        'name': client.name,
                        'client_code': client.client_code,
                        'client_type': client.client_type
                    })

                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    continue

            return Response({
                'message': f'Excel upload completed',
                'created_count': len(created_clients),
                'error_count': len(errors),
                'created_clients': created_clients,
                'errors': errors
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Failed to process Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """Download Excel template for client upload"""
        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Client Template"

        # Define headers
        headers = [
            'Name', 'Client Code', 'Client Type', 'GSTIN', 'PAN',
            'Email', 'Phone', 'State Code', 'Credit Terms (Days)',
            'Billing Address Line 1', 'Billing Address Line 2',
            'Billing City', 'Billing State', 'Billing Pincode',
            'Shipping Address Line 1', 'Shipping Address Line 2',
            'Shipping City', 'Shipping State', 'Shipping Pincode'
        ]

        # Add headers to first row
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)

        # Add sample data
        sample_data = [
            ['ABC Corporation', 'CLI001', 'b2b', '29AABCU9603R1ZX', 'AABCU9603R', 'abc@corp.com', '9876543210', '29', '30',
             '123 Business Street', 'Tower A', 'Mumbai', 'Maharashtra', '400001',
             '123 Business Street', 'Tower A', 'Mumbai', 'Maharashtra', '400001'],
            ['John Doe', 'CLI002', 'b2c', '', '', 'john@email.com', '9876543211', '07', '15',
             '456 Residential Area', '', 'Delhi', 'Delhi', '110001',
             '456 Residential Area', '', 'Delhi', 'Delhi', '110001']
        ]

        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx).value = value

        # Add notes in a separate sheet
        notes_sheet = workbook.create_sheet("Instructions")
        instructions = [
            "Instructions for Client Upload:",
            "",
            "1. Required Fields (must be filled):",
            "   - Name: Client/Company name",
            "   - Client Code: Unique identifier for the client",
            "   - Client Type: Either 'b2b' or 'b2c'",
            "",
            "2. Optional Fields:",
            "   - GSTIN: 15-digit GST identification number",
            "   - PAN: 10-digit permanent account number",
            "   - Email: Contact email address",
            "   - Phone: Contact phone number",
            "   - State Code: 2-digit state code (01-37)",
            "   - Credit Terms: Payment terms in days (default: 30)",
            "",
            "3. Address Fields:",
            "   - Both billing and shipping addresses are optional",
            "   - Fill as many fields as available for better records",
            "",
            "4. Important Notes:",
            "   - Client Code must be unique within your tenant",
            "   - Client Type must be exactly 'b2b' or 'b2c'",
            "   - GSTIN format: 2-digit state code + 10-digit PAN + 1-digit entity code + 1-digit checksum + 1-letter",
            "   - Remove sample data before uploading your actual client data",
            "",
            "5. Sample Data:",
            "   - Two sample rows are provided as examples",
            "   - Delete these rows and add your actual client data"
        ]

        for row_idx, instruction in enumerate(instructions, 1):
            notes_sheet.cell(row=row_idx, column=1).value = instruction

        # Save to BytesIO
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        # Create HTTP response
        response = HttpResponse(
            file_stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="client_upload_template.xlsx"'

        return response

    @action(detail=False, methods=['get'], url_path='export')
    def export_clients(self, request):
        """Export all clients to Excel"""
        user = request.user
        tenant_ids = user.tenant_memberships.filter(is_active=True).values_list('tenant_id', flat=True)
        clients = Client.objects.filter(tenant_id__in=tenant_ids).order_by('name')

        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Clients Export"

        # Headers
        headers = [
            'Name', 'Client Code', 'Client Type', 'GSTIN', 'PAN',
            'Email', 'Phone', 'State Code', 'Credit Terms (Days)',
            'Billing Address', 'Shipping Address', 'Created Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for row_idx, client in enumerate(clients, 2):
            # Format addresses
            billing_addr = client.billing_address or {}
            shipping_addr = client.shipping_address or {}

            billing_str = ', '.join([
                billing_addr.get('line1', ''),
                billing_addr.get('line2', ''),
                billing_addr.get('city', ''),
                billing_addr.get('state', ''),
                billing_addr.get('pincode', '')
            ]).strip(', ')

            shipping_str = ', '.join([
                shipping_addr.get('line1', ''),
                shipping_addr.get('line2', ''),
                shipping_addr.get('city', ''),
                shipping_addr.get('state', ''),
                shipping_addr.get('pincode', '')
            ]).strip(', ')

            data = [
                client.name,
                client.client_code,
                client.client_type.upper(),
                client.gstin or '',
                client.pan or '',
                client.email or '',
                client.phone or '',
                client.state_code or '',
                client.credit_terms_days,
                billing_str,
                shipping_str,
                client.created_at.strftime('%Y-%m-%d')
            ]

            for col_idx, value in enumerate(data, 1):
                worksheet.cell(row=row_idx, column=col_idx).value = value

        # Save to BytesIO
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        # Create HTTP response
        response = HttpResponse(
            file_stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="clients_export.xlsx"'

        return response


class TenantMembershipViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = TenantMembershipSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # Users can only see memberships from their tenants
        tenant_ids = user.tenant_memberships.filter(is_active=True).values_list('tenant_id', flat=True)
        return TenantMembership.objects.filter(tenant_id__in=tenant_ids, is_active=True)

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        membership = self.get_object()
        user = request.user

        # Check if user has permission to deactivate this membership
        user_membership = TenantMembership.objects.filter(
            user=user,
            tenant=membership.tenant,
            is_active=True,
            role__in=['owner', 'admin']
        ).first()

        if not user_membership:
            return Response(
                {'error': 'You do not have permission to deactivate this membership'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Prevent deactivating the last owner
        if membership.role == 'owner':
            owner_count = TenantMembership.objects.filter(
                tenant=membership.tenant,
                role='ca_owner',
                is_active=True
            ).count()
            if owner_count <= 1:
                return Response(
                    {'error': 'Cannot deactivate the last owner of the tenant'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        membership.is_active = False
        membership.save()

        return Response({'message': 'Membership deactivated successfully'})


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AuditLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['entity_type', 'action', 'is_financial_transaction', 'compliance_relevant']
    search_fields = ['entity_type', 'entity_id', 'action']
    ordering = ['-timestamp']

    def get_queryset(self):
        user = self.request.user
        # Users can only see audit logs from their tenants
        tenant_ids = user.tenant_memberships.filter(is_active=True).values_list('tenant_id', flat=True)
        return AuditLog.objects.filter(tenant_id__in=tenant_ids)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def current_tenant(request):
    """Get current user's primary tenant"""
    membership = request.user.tenant_memberships.filter(is_active=True).first()
    if not membership:
        return Response(
            {'error': 'User is not a member of any tenant'},
            status=status.HTTP_404_NOT_FOUND
        )

    serializer = TenantSerializer(membership.tenant)
    return Response({
        'tenant': serializer.data,
        'membership': TenantMembershipSerializer(membership).data
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def tenant_setup(request):
    """Setup a new tenant for the user"""
    user = request.user

    # Check if user already has an active tenant
    if user.tenant_memberships.filter(is_active=True).exists():
        return Response(
            {'error': 'User already belongs to a tenant'},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = TenantSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        tenant = serializer.save()
        membership = TenantMembership.objects.get(user=user, tenant=tenant)

        return Response({
            'tenant': TenantSerializer(tenant).data,
            'membership': TenantMembershipSerializer(membership).data,
            'message': 'Tenant setup completed successfully'
        }, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)